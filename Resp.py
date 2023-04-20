import argparse
import random

from openpyxl import load_workbook

argParser = argparse.ArgumentParser()
argParser.add_argument("-m", "--month", help="choose month")
argParser.add_argument("-w", "--workers", type=int, default=17, help="input nr of workers")
args = argParser.parse_args()

# Wybierz skoroszyt i arkusz
wb = load_workbook('Grafik.xlsx')
ws = wb[args.month]

# Zacznij od I2
row = 2
col = 8
nr_of_workers = args.workers

# Lista gdy tylko 2 osoby na zmianie
listShort = [' - UnCQ', ' - R']

# Lista gdy 3+ osób na zmianie
listFull = [' - Un', ' - CQ', ' - R']


# Przywróć do stanu pierwotnego
def destroy():
    global row, col
    while ws[1][col].value:
        row = 2
        while row <= nr_of_workers:
            if ws[row][col].value != 'w' and ws[row][col].value != 'w!' and ws[row][col].value != 'u' and \
                    ws[row][col].value != '\xa0' and ws[row][col].value is not None:
                ws[row][col].value = int(str(ws[row][col].value)[0])
            row = row + 1
        col = col + 1
    row = 2
    col = 8


# Stwórz klona dla każdej listy
def list_clone():
    global list0s, list0f, list1s, list1f, list2s, list2f
    list0s = listShort[:]
    list0f = listFull[:]
    list1s = listShort[:]
    list1f = listFull[:]
    list2s = listShort[:]
    list2f = listFull[:]


# Sprawdź ile zmian danego dnia
def countPeople(x, y):
    global count0, count1, count2

    # Dla 17 wierszy
    while x <= nr_of_workers:

        if isinstance(ws[x][y].value, str):
            ws[x][y].value = ws[x][y].value

        if isinstance(ws[x][y].value, int):
            if ws[x][y].value == 0:
                count0 = count0 + 1
            if ws[x][y].value == 1:
                count1 = count1 + 1
            if ws[x][y].value == 2:
                count2 = count2 + 1
        x = x + 1

        extend_list()


# Dodaj wystarczającą ilośc pustych
def extend_list():
    while len(list0f) < count0:
        list0f.append('')

    while len(list1f) < count1:
        list1f.append('')

    while len(list2f) < count2:
        list2f.append('')


# Komórka ma 0
def resp0():
    # Jeżeli tylko 2, skorzystaj z shortList
    if count0 == 2:
        random.shuffle(list0s)

        # Jeżeli obecna komórka i element z wierzchu listy, nie dają takiej samej wartości co wartość poprzedniego dnia, to dopisz ją
        if str(ws[row][col].value) + list0s[-1] != str(ws[row][col - 1].value):
            ws[row][col].value = str(ws[row][col].value) + list0s.pop()

        # W przeciwnym wypadku daj element z drugiej strony (może się okazać że w liście jest tylko jeden element, ale w ten sposób ograniczam duplikaty i błędy
        else:
            ws[row][col].value = str(ws[row][col].value) + list0s.pop(0)
    if count0 > 2:
        random.shuffle(list0f)
        if str(ws[row][col].value) + list0f[-1] != str(ws[row][col - 1].value):
            ws[row][col].value = str(ws[row][col].value) + list0f.pop()
        else:
            ws[row][col].value = str(ws[row][col].value) + list0f.pop(0)


def resp1():
    if count1 == 2:
        random.shuffle(list1s)
        if str(ws[row][col].value) + list1s[-1] != str(ws[row][col - 1].value):
            ws[row][col].value = str(ws[row][col].value) + list1s.pop()
        else:
            ws[row][col].value = str(ws[row][col].value) + list1s.pop(0)
    elif count1 == 3:
        random.shuffle(list1f)
        if str(ws[row][col].value) + list1f[-1] != str(ws[row][col - 1].value):
            ws[row][col].value = str(ws[row][col].value) + list1f.pop()
        else:
            ws[row][col].value = str(ws[row][col].value) + list1f.pop(0)

    # Staraj się przypisać Kamilowi i Adrianowi puste
    elif count1 >= 4 and ws[row][0].value == 'Piotr':
        ws[row][col].value = str(ws[row][col].value) + list1f.pop(list1f.index(''))
    elif count1 >= 4 and ws[row][0].value == 'Adrian':
        ws[row][col].value = str(ws[row][col].value) + list1f.pop()
    else:
        random.shuffle(list1f)
        default = str(ws[row][col].value) + str(list1f[-1])
        previous = str(ws[row][col - 1].value)

        # To przypisanie jest przydatne, ale z jakiegoś powodu wsadziłem to w dziwnym miejscu ;p
        if default != previous:
            ws[row][col].value = str(ws[row][col].value) + list1f.pop(-1)
        else:
            ws[row][col].value = str(ws[row][col].value) + list1f.pop(0)


def resp2():
    if count2 == 2:
        random.shuffle(list2s)
        if str(ws[row][col].value) + list2s[-1] != str(ws[row][col - 1].value):
            ws[row][col].value = str(ws[row][col].value) + list2s.pop()
        else:
            ws[row][col].value = str(ws[row][col].value) + list2s.pop(0)
    elif count2 == 3:
        random.shuffle(list2f)
        if str(ws[row][col].value) + list2f[-1] != str(ws[row][col - 1].value):
            ws[row][col].value = str(ws[row][col].value) + list2f.pop()
        else:
            ws[row][col].value = str(ws[row][col].value) + list2f.pop(0)
    elif count2 >= 4 and ws[row][0].value == 'Piotr':
        ws[row][col].value = str(ws[row][col].value) + list2f.pop()
    elif count2 >= 4 and ws[row][0].value == 'Adrian':
        ws[row][col].value = str(ws[row][col].value) + list2f.pop()
    else:
        random.shuffle(list2f)
        default = str(ws[row][col].value) + str(list2f[-1])
        previous = str(ws[row][col - 1].value)

        if default != previous:
            ws[row][col].value = str(ws[row][col].value) + list2f.pop(-1)
        else:
            ws[row][col].value = str(ws[row][col].value) + list2f.pop(0)


if __name__ == '__main__':
    destroy()

    while ws[1][col].value:
        # while b < 18:
        list_clone()

        count0 = 0
        count1 = 0
        count2 = 0

        row = 2

        countPeople(row, col)

        while row <= nr_of_workers:

            if ws[row][col].value == 0:
                resp0()
            elif ws[row][col].value == 1:
                resp1()
            elif ws[row][col].value == 2:
                resp2()

            # Tutaj, powinno w 20 wierszu pokazywać czy znalazł duplikat
            if ws[row][col].value == ws[row][col - 1].value and \
                    len(str(ws[row][col].value)) > 1 and ws[row][col].value != 'w!':
                ws[20][col].value = 'Duplikat'
                ws[21][col].value = ws[row][col].coordinate

            # Zamień te zmiany które dostały puste na int. Inaczej w excelu komorki wskazują błąd
            if len(str(ws[row][col].value)) == 1 and ws[row][col].value != 'w' and ws[row][col].value != 'u' \
                    and ws[row][col].value != '\xa0' and ws[row][col].value is not None:
                ws[row][col].value = int(ws[row][col].value)

                # ws[19][b].value = 'Duplikat'
            row = row + 1
        col = col + 1

    wb.save('Grafik.xlsx')
